import openpyxl
from cbr_exchange import get_rates, get_keyrate
from daily import update_daily_sheet_core, update_cash_in_bank_core
from helper import get_filename, date_extract, date_fallback, find_excel_file_in_current_dir, file_save
# , clear_external_reference
from table import table_new_column, copy_cpfo, copy_apk, copy_rbpi, copy_severnaya, copy_woysk, copy_stesha
from datetime import datetime
import sys

# magic const
USD = 'USD/RUB'
EUR = 'EUR/RUB'
CNY = 'CNY/RUB'
THB = 'THB/RUB'
SHEET_1 = 'Daily'
SHEET_CASH_IN_BANK = 'Cash in bank report'
SHEET_TABLE = 'Table'
SHEET_DAILY_EXCHANGE = 'Daily exchange'
DEFAULT_KEYRATE = 18.0
PREFIX_CPFO = "Cash report_"
PREFIX_APK = "APK DON Deposit&loan "
PREFIX_RBPI = "RBPI DepositLoan Weekly report "
PREFIX_SEV = "Копия Cash_Severna"
PREFIX_WOYSK = "Financial memorandum SW_"
PREFIX_STESHA = "Stesha Cash report_"

filename = get_filename(PREFIX_CPFO)
date_str = date_extract(filename)
if date_str is None:
    date_str = date_fallback()
KR_date = datetime.strptime(date_str, '%d/%m/%Y').date()
currency = get_rates(date_str)

keyrate = get_keyrate(KR_date) or DEFAULT_KEYRATE
source_workbooks = {}
try:
    print("\n--- Загрузка исходных файлов Excel ---")
        
    # Основной файл
    main_excel_path = find_excel_file_in_current_dir()
    wb_formulas = openpyxl.load_workbook(main_excel_path)
    # Второй раз грузим тот же файл для получения вычисленных значений
    wb_values = openpyxl.load_workbook(main_excel_path, data_only=True)

    # Загружаем все остальные файлы один раз
    source_files_map = {
        'cpfo': get_filename(PREFIX_CPFO),
        'apk': get_filename(PREFIX_APK),
        'rbpi': get_filename(PREFIX_RBPI),
        'severnaya': get_filename(PREFIX_SEV),
        'woysk': get_filename(PREFIX_WOYSK),
        'stesha': get_filename(PREFIX_STESHA)
    }

    for key, path in source_files_map.items():
        if path:
            print(f"Загружаю '{path}'...")
            source_workbooks[key] = openpyxl.load_workbook(path, data_only=True)
        else:
            print(f"Предупреждение: не найден исходный файл для '{key}'.")
            source_workbooks[key] = None
except FileNotFoundError as e:
        print(f"Критическая ошибка: {e}. Работа программы прекращена.")
        sys.exit(1)
except Exception as e:
        print(f"Неожиданная ошибка при загрузке файлов: {e}. Работа программы прекращена.")
        sys.exit(1)

print("--- Все файлы успешно загружены в память. Начинаю обработку. ---\n")
try:
    # clear_external_reference(wb_formulas, wb_values)
    sums = update_daily_sheet_core(
        wb_formulas,
        wb_values,
        SHEET_1,
        KR_date,
        keyrate,
        currency[USD],
        currency[EUR],
        currency[CNY]
    )
    update_cash_in_bank_core(wb_formulas, SHEET_CASH_IN_BANK, KR_date, currency[THB])
    column = table_new_column(wb_formulas, wb_values, SHEET_TABLE, KR_date)
    if source_workbooks.get('cpfo'):
        copy_cpfo(wb_formulas, source_workbooks['cpfo'], column, SHEET_TABLE)
    
    if source_workbooks.get('apk'):
        copy_apk(wb_formulas, source_workbooks['apk'], column, SHEET_TABLE)
        
    if source_workbooks.get('rbpi'):
        copy_rbpi(wb_formulas, source_workbooks['rbpi'], column, SHEET_TABLE)

    deposit_sev = copy_severnaya(wb_formulas, source_workbooks.get('severnaya'), column, SHEET_CASH_IN_BANK, SHEET_TABLE, SHEET_DAILY_EXCHANGE, KR_date)
    deposit_woysk = copy_woysk(wb_formulas, source_workbooks.get('woysk'), column, SHEET_TABLE)
    deposit_stesha = copy_stesha(wb_formulas, source_workbooks.get('stesha'), column, SHEET_CASH_IN_BANK, SHEET_TABLE)
    print(f"\n\n------------------П-Р-О-В-Е-Р-К-А---Д-Е-П-О-З-И-Т-О-В------------------\n\n")
    if sums is None:
        print("Суммы депозитов не найдены.")
    else:
        if deposit_sev != sums['sev']:
            print("Надо обновить депозиты Северной.")
        else:
            print("Не нужно обновлять депозиты Северной.")
        if deposit_woysk != sums['woysk']:
            print("Надо обновить депозиты Войсковиц.")
        else:
            print("Не нужно обновлять депозиты Войсковиц.")
        if deposit_stesha != sums['stesha']:
            print("Надо обновить депозиты Стеши.")
        else:
            print("Не нужно обновлять депозиты Стеши.")
    file_save(main_excel_path, KR_date, wb_formulas)
except Exception as e:
    print(f"An unexpected error occurred during the process: {e}")
finally:
    print("\n--- Закрытие рабочих книг ---")
    wb_formulas.close()
    wb_values.close()
    for wb in source_workbooks.values():
        if wb:
            wb.close()
    print("--- Обработка завершена ---")    