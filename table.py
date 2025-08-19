import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string
from helper import copy_cell_style, get_filename, divide, clean_and_convert_to_float, update_formula_and_compare

def table_new_column(wb_formulas, wb_values, sheet_name, report_date):
    if sheet_name not in wb_formulas.sheetnames:
        print(f"Лист с именем '{sheet_name}' не найден.")
        return None
           
    ws = wb_formulas[sheet_name]
    sheet_values = wb_values[sheet_name]
    print(f"Работаю с листом {sheet_name}...")

    last_col_idx = ws.max_column
    new_col_idx = last_col_idx + 1
    new_column_name = get_column_letter(new_col_idx)
    print(f"Последний столбец с данными: {get_column_letter(last_col_idx)}. Новый столбец: {new_column_name}.")
    
    for row_num in range(1, 51):
        source_cell = ws.cell(row=row_num, column=last_col_idx)
        dest_cell = ws.cell(row=row_num, column=new_col_idx)
        
        source_cell_static_value = sheet_values.cell(row=row_num, column=last_col_idx).value

        copy_cell_style(source_cell, dest_cell)

        # Временно для сипи трейдинга
        if row_num == 47:
            dest_cell.value = source_cell_static_value
        
        elif source_cell.data_type == 'f':
            formula = source_cell.value
            translator = Translator(formula, origin=source_cell.coordinate)
            dest_cell.value = translator.translate_formula(dest_cell.coordinate)

            if source_cell_static_value is not None:
                source_cell.value = source_cell_static_value
        
    date_row = 1
    date_dest_cell = ws.cell(row=date_row, column=new_col_idx)
    date_dest_cell.value = report_date
    print("Готово!")
    return new_column_name



def copy_cpfo(wb_formulas, column, sheet_name):
    source_filename = get_filename(fixed_part = "Cash report_")
    try:
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        source_ws = source_wb.active
        if source_ws is None:
            print(f"Ошибка: Не удалось найти активный лист в файле '{source_filename}'.")
            return
        source_range = 'B3:G3'
        target_ws = wb_formulas[sheet_name]
        source_values_raw = [cell.value for cell in source_ws[source_range][0]]
        if all(v is None for v in source_values_raw):
            print("Значения не найдены")
            return
        print(f"Прочитанные значения: {source_values_raw}")
        processed_values = [divide(val) for val in source_values_raw]
        print(f"Обработанные значения (в млн): {processed_values}")
        start_row_idx = 5
        start_col_idx = column_index_from_string(column)
        for i, value in enumerate(processed_values):
            target_ws.cell(row=start_row_idx + i, column=start_col_idx, value=value)
        print(f"Данные успешно скопированы в столбец {column} листа '{sheet_name}'.")
    except FileNotFoundError:
        print(f"Файл {source_filename} не найден.")
    
def copy_apk(wb_formulas, column, sheet_name):
    source_filename = get_filename(fixed_part = "APK DON Deposit&loan ")

    try:
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        source_ws = source_wb.active
        
        if source_ws is None:
            print(f"Ошибка: Не найден активный лист в '{source_filename}'.")
            return
            
        source_range = 'E3:O3'
        target_ws = wb_formulas[sheet_name]

        # Читаем значения как есть, без обработки
        values_to_copy = [cell.value for cell in source_ws[source_range][0]]
        print(f"Прочитанные значения из {source_filename}: {values_to_copy}")

        if all(v is None for v in values_to_copy):
            print(f"Значения в '{source_filename}' не найдены. Операция прервана.")
            return

        # Записываем значения в целевой файл
        start_row_idx = 13
        start_col_idx = column_index_from_string(column)

        for i, value in enumerate(values_to_copy):
            target_ws.cell(row=start_row_idx + i, column=start_col_idx, value=value)

        print(f"Данные из {source_filename} успешно скопированы в столбец {column} (строки 13-23).")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_filename}' не найден.")
    except Exception as e:
        print(f"Не удалось обработать файл '{source_filename}'. Ошибка: {e}")

def copy_rbpi(wb_formulas, column, sheet_name):
    source_filename = get_filename(fixed_part = "RBPI DepositLoan Weekly report ")

    try:
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        source_ws = source_wb.active

        if source_ws is None:
            print(f"Ошибка: Не найден активный лист в '{source_filename}'.")
            return

        source_range = 'E3:R3'
        target_ws = wb_formulas[sheet_name]

        # Читаем значения как есть, без обработки
        values_to_copy = [cell.value for cell in source_ws[source_range][0]]
        print(f"Прочитанные значения из {source_filename}: {values_to_copy}")

        if all(v is None for v in values_to_copy):
            print(f"Значения в '{source_filename}' не найдены. Операция прервана.")
            return

        # Записываем значения в целевой файл
        start_row_idx = 27
        start_col_idx = column_index_from_string(column)

        for i, value in enumerate(values_to_copy):
            target_ws.cell(row=start_row_idx + i, column=start_col_idx, value=value)

        print(f"Данные из {source_filename} успешно скопированы в столбец {column} (строки 27-40).")
    
    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_filename}' не найден.")
    except Exception as e:
        print(f"Не удалось обработать файл '{source_filename}'. Ошибка: {e}")

def copy_severnaya(wb_formulas, column, sheet_daily, sheet_cib_name, sheet_table_name, sheet_de_name, report_date):
    source_filename = get_filename(fixed_part = "Cash_Severna")

    try:
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        
        sheet_to_process = "Текущие счета"
        if sheet_to_process not in source_wb.sheetnames:
            print(f"Ошибка: Лист '{sheet_to_process}' не найден в файле '{source_filename}'.")
            return
        
        ws = source_wb[sheet_to_process]

        last_data_col_idx = None
        for col_idx in range(ws.max_column, 3, -1):
            cell_value = ws.cell(row=4, column=col_idx).value
            # Ищем первую непустую ячейку
            if cell_value is not None and cell_value != '':
                prev_1_is_empty = ws.cell(row=4, column=col_idx - 1).value in (None, '')
                prev_2_is_empty = ws.cell(row=4, column=col_idx - 2).value in (None, '')
                prev_3_is_empty = ws.cell(row=4, column=col_idx - 3).value in (None, '')
                if prev_1_is_empty and prev_2_is_empty and prev_3_is_empty:
                    print(f"Столбец {get_column_letter(col_idx)}, не последний, поиск продолжается...")
                    continue
                else:
                    last_data_col_idx = col_idx
                    print(f"Найдена последняя колонка с данными в '{source_filename}': {get_column_letter(last_data_col_idx)}")
                    break # Выходим из цикла, как только нашли

        if last_data_col_idx is None:
            print(f"Не удалось найти колонку с данными в '{source_filename}' на листе '{sheet_to_process}'.")
            return

         # 2. Расчет всех сумм
        sum_ranges = {
            'rub': (4, 13), 'eur': (14, 23), 'usd': (23, 31), 'cny': (31, 42)
        }
        sums = {}
        for key, (start, end) in sum_ranges.items():
            current_sum = sum(clean_and_convert_to_float(ws.cell(row, last_data_col_idx).value) for row in range(start, end))
            sums[key] = current_sum / 1_000_000 # Сразу делим на миллион
        
        # 3. Вставка суммы RUB в 'Table'
        wb_formulas[sheet_table_name].cell(row=45, column=column_index_from_string(column), value=sums['rub'])
        print(f"  - Сумма RUB ({sums['rub']:.2f}) записана в '{sheet_table_name}'.")

        # 4. Обновление формул и сравнение
        target_ws_cib = wb_formulas[sheet_cib_name]
        is_new_row_needed = False

        # (адрес ячейки, новое значение, название валюты)
        updates_to_perform = [
            ('E52', sums['eur'], 'EUR'),
            ('E51', sums['usd'], 'USD'),
            ('E53', sums['cny'], 'CNY'),
        ]

        for address, new_val, name in updates_to_perform:
            if update_formula_and_compare(target_ws_cib, address, new_val, name):
                is_new_row_needed = True # Если хоть одно сравнение True, поднимаем флаг

        # 5. Условное добавление строки в 'Daily exchange'
        if is_new_row_needed:
            print("Обнаружено превышение старых значений, добавляю новую строку в 'Daily exchange'...")
            target_ws_de = wb_formulas[sheet_de_name]
            last_row = target_ws_de.max_row
            new_row = last_row + 1
            if last_row in target_ws_de.row_dimensions:
                target_ws_de.row_dimensions[new_row].height = target_ws_de.row_dimensions[last_row].height
            # Копирование стиля со всех ячеек предыдущей строки
            for col in range(1, target_ws_de.max_column + 1):
                copy_cell_style(target_ws_de.cell(last_row, col), target_ws_de.cell(new_row, col))
            
            # Вставка даты
            target_ws_de.cell(new_row, 2).value = report_date # столбец B
            print(f"ВНИМАНИЕ: На лист '{sheet_de_name}' добавлена новая строка ({new_row}) с датой {report_date.strftime('%d.%m.%Y')}.")
        else:
            print("Превышений над старыми значениями не обнаружено, новая строка не требуется.")
        # ДЕПОЗИТЫ
        ws_deposits = source_wb["Депозиты"]
        
        total_rur_row = None
        for r_idx in range(1, ws_deposits.max_row + 1):
            if ws_deposits.cell(row=r_idx, column=2).value == "Total RUR":
                total_rur_row = r_idx
                break
        
        if not total_rur_row:
            print("  - ОШИБКА: Не найдена строка 'Total RUR' на листе 'Депозиты'. Операция прервана.")
            return

        # --- ШАГ 2: СБОР ДАННЫХ О ДЕПОЗИТАХ ---
        deposit_data = []
        for r_idx in range(8, total_rur_row):
            # Считаем строку валидной, если в столбце D есть значение
            if ws_deposits.cell(row=r_idx, column=4).value not in (None, ''):
                deposit_data.append({
                    'start_date': ws_deposits.cell(row=r_idx, column=5).value, # E -> A
                    'due_date':   ws_deposits.cell(row=r_idx, column=6).value, # F -> B
                    'rate':       ws_deposits.cell(row=r_idx, column=7).value, # G -> C
                    'amount':     ws_deposits.cell(row=r_idx, column=4).value, # A -> D
                })
        
        num_deposits = len(deposit_data)
        print(f"  - Найдено {num_deposits} депозитов для синхронизации.")

        # --- ШАГ 3: ПРОВЕРКА ДАННЫХ ПОСЛЕ "Total RUR" ---
        has_data_after = any(
            ws_deposits.cell(row=r_idx, column=c_idx).value not in (None, '')
            for r_idx in range(total_rur_row + 1, min(total_rur_row + 11, ws_deposits.max_row + 1))
            for c_idx in range(1, 8)
        )
        if has_data_after:
            print("  - ВНИМАНИЕ: Обнаружены непустые строки в пределах 10 строк после 'Total RUR'. Проверьте исходный файл.")

        # --- ШАГ 4: АНАЛИЗ ЦЕЛЕВОГО ЛИСТА "Daily" ---
        ws_daily = wb_formulas[sheet_daily]
        start_row_daily = 34
        first_empty_daily_row = start_row_daily
        while ws_daily.cell(row=first_empty_daily_row, column=1).value is not None:
            first_empty_daily_row += 1
        num_daily_rows = first_empty_daily_row - start_row_daily
        
        # --- ШАГ 5: СИНХРОНИЗАЦИЯ КОЛИЧЕСТВА СТРОК ---
        diff = num_deposits - num_daily_rows
        if diff > 0:
            ws_daily.insert_rows(first_empty_daily_row, amount=diff)
            print(f"  - Вставлено {diff} строк(и) в лист 'Daily'.")
        elif diff < 0:
            start_delete_row = first_empty_daily_row - abs(diff)
            ws_daily.delete_rows(start_delete_row, amount=abs(diff))
            print(f"  - Удалено {abs(diff)} строк(и) из листа 'Daily'.")
        else:
            print("  - Количество строк в 'Daily' совпадает, модификация структуры не требуется.")

        new_last_row_nn = start_row_daily + num_deposits - 1

        # --- ШАГ 6: КОПИРОВАНИЕ ДАННЫХ ---
        for i, data in enumerate(deposit_data):
            target_row = start_row_daily + i
            ws_daily.cell(target_row, 1).value = data['start_date']
            ws_daily.cell(target_row, 2).value = data['due_date']
            ws_daily.cell(target_row, 3).value = data['rate']
            amount_rub = clean_and_convert_to_float(data['amount'])
            ws_daily.cell(target_row, 4).value = divide(amount_rub) if amount_rub else 0
        print("  - Данные по депозитам успешно скопированы в 'Daily'.")

        # --- ШАГ 7: ОБНОВЛЕНИЕ ФОРМУЛЫ В "Cash in bank report" ---
        ws_cib = wb_formulas[sheet_cib_name]
        new_formula = f"=SUM(Daily!D{start_row_daily}:D{new_last_row_nn})"
        ws_cib['E47'].value = new_formula
        print(f"  - Формула в 'Cash in bank report'!E47 обновлена на: {new_formula}")
        print("Синхронизация депозитов завершена.")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_filename}' не найден.")
    except Exception as e:
        print(f"Не удалось обработать файл '{source_filename}'. Ошибка: {e}")

def copy_woysk(wb_formulas, column, sheet_name):
    source_filename = get_filename(fixed_part = "Financial memorandum SW_")

    try:
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        
        sheet_to_process = "accounts"
        if sheet_to_process not in source_wb.sheetnames:
            print(f"Ошибка: Лист '{sheet_to_process}' не найден в файле '{source_filename}'.")
            return
        
        ws = source_wb[sheet_to_process]

        last_data_col_idx = None
        for col_idx in range(ws.max_column, 0, -1):
            cell_value = ws.cell(row=32, column=col_idx).value
            # Ищем первую непустую ячейку
            if cell_value is not None and cell_value != '':
                last_data_col_idx = col_idx
                print(f"Найдена последняя колонка с данными в '{source_filename}': {get_column_letter(last_data_col_idx)}")
                break # Выходим из цикла, как только нашли

        if last_data_col_idx is None:
            print(f"Не удалось найти колонку с данными в '{source_filename}' на листе '{sheet_to_process}'.")
            return

        # 2. Суммируем значения в найденном столбце
        total_sum = 0.0
        for row_idx in range(32, 39):
            cell_value = ws.cell(row=row_idx, column=last_data_col_idx).value
            total_sum += clean_and_convert_to_float(cell_value)
        
        print(f"Сумма по столбцу {get_column_letter(last_data_col_idx)} (строки 32-38): {total_sum:,.2f}")

        # 3. Делим на 1 миллион
        final_value = divide(total_sum)
        print(f"Итоговое значение для вставки (сумма / 1 млн): {final_value}")

        # 4. Вставляем в целевую книгу
        target_ws = wb_formulas[sheet_name]
        target_row = 46
        target_col_idx = column_index_from_string(column)
        
        target_ws.cell(row=target_row, column=target_col_idx, value=final_value)
        
        print(f"Значение успешно записано в столбец {column}, строку {target_row}.")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_filename}' не найден.")
    except Exception as e:
        print(f"Не удалось обработать файл '{source_filename}'. Ошибка: {e}")

def copy_stesha(wb_formulas, column, target_sheet_name, sheet_name):
    source_filename = get_filename(fixed_part = "Stesha Cash report_")

    try:
        # Загружаем книгу только со значениями
        source_wb = openpyxl.load_workbook(source_filename, data_only=True)
        
        sheet1 = "Cash in bank report"
        if sheet1 not in source_wb.sheetnames:
            print(f"Ошибка: Лист '{sheet1}' не найден в файле '{source_filename}'.")
            return
        
        ws = source_wb[sheet1]

        # 1. Ищем последнюю непустую строку в столбце B (индекс 2), двигаясь снизу вверх
        last_value_raw = None
        last_row_found = None
        # Идем от последней строки листа вверх к первой
        for row_idx in range(ws.max_row, 0, -1):
            cell_value = ws.cell(row=row_idx, column=2).value # Столбец B
            # Ищем первую непустую ячейку
            if cell_value is not None and str(cell_value).strip() != '':
                last_value_raw = cell_value
                last_row_found = row_idx
                print(f"Найдено последнее значение в '{source_filename}': '{last_value_raw}' в строке {last_row_found}")
                break # Выходим из цикла, как только нашли

        if last_value_raw is None:
            print(f"Не удалось найти данные в столбце B файла '{source_filename}'.")
            return

        # 2. Очищаем и преобразуем значение в число
        final_value = clean_and_convert_to_float(last_value_raw)
        print(f"Итоговое значение для вставки: {final_value}")

        # 3. Вставляем в целевую книгу
        target_ws = wb_formulas[sheet_name]
        target_row = 48
        target_col_idx = column_index_from_string(column)
        
        target_ws.cell(row=target_row, column=target_col_idx, value=final_value)
        
        print(f"Значение успешно записано в столбец {column}, строку {target_row}.")
        # Обновление листа Cash in bank report (валюты)
        sheet2_name = "Daily exchange"
        if sheet2_name in source_wb.sheetnames:
            ws2 = source_wb[sheet2_name]
            last_value_i = None
            for row_idx in range(ws2.max_row, 0, -1):
                cell_value = ws2.cell(row=row_idx, column=9).value
                if cell_value is not None and str(cell_value).strip() != '':
                    last_value_i = cell_value
                    break

            if last_value_i is not None:
                new_variable_part = clean_and_convert_to_float(last_value_i)
                if target_sheet_name in wb_formulas.sheetnames:
                    target_ws_cib = wb_formulas[target_sheet_name]
                    target_cell = target_ws_cib['G53']
                    existing_formula = target_cell.value

                    if existing_formula and str(existing_formula).startswith('='):
                        parts = str(existing_formula).split('*', 1)
                        if len(parts) > 1:
                            static_part = parts[1]
                            new_formula = f"={new_variable_part}*{static_part}"
                            target_cell.value = new_formula
                            print(f"Формула в '{target_sheet_name}'!G53 успешно обновлена на: {new_formula}")
                        else:
                            print(f"Ошибка - формула в G53 имеет неожиданный формат.")
                    else:
                        print(f"Задача 2: Ошибка - ячейка G53 не содержит формулу.")
                else:
                    print(f"Задача 2: Ошибка - целевой лист '{target_sheet_name}' не найден.")
            else:
                print(f"Задача 2: Не найдены данные в ст. I на листе '{sheet2_name}'.")
        else:
            print(f"Задача 2: Лист '{sheet2_name}' не найден в '{source_filename}'.")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{source_filename}' не найден.")
    except Exception as e:
        print(f"Не удалось обработать файл '{source_filename}'. Ошибка: {e}")