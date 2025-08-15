from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from helper import copy_cell_style

def find_anchor_column(sheet, anchor_text="Rate from CBR"):
    for row in sheet.iter_rows(max_row=10):
        for cell in row:
            if cell.value and anchor_text in str(cell.value):
                print(f"Работаю со столбцом {cell.column}.")
                return cell.column
    return None

def update_daily_sheet_core(
    wb_formulas,
    wb_values,
    sheet_name,
    new_date,
    new_key_rate,
    new_rub_usd,
    new_rub_eur,
    new_rub_cny
    ):
    
    if sheet_name not in wb_formulas.sheetnames:
        print(f"Лист с именем '{sheet_name}' не найден.")
        return None
    sheet_formulas = wb_formulas[sheet_name]
    
    sheet_values = wb_values[sheet_name]

    anchor_col_index = find_anchor_column(sheet_values)
    if not anchor_col_index:
        print(f"Текст 'Rate from CBR' не найден на листе.")
        return None

    new_column_index = anchor_col_index 
    previous_column_index = new_column_index - 1
    
    print(f"Последний столбец с данными: {get_column_letter(previous_column_index)}. Новый столбец: {get_column_letter(new_column_index)}")

    sheet_formulas.insert_cols(new_column_index)
    print("Вставил новый столбец. Работаю со строками 3-30...")
    
    for row_idx in range(3, 31):
        previous_cell_formulas = sheet_formulas.cell(row=row_idx, column=previous_column_index)
        new_cell_formulas = sheet_formulas.cell(row=row_idx, column=new_column_index)

        previous_cell_static_value = sheet_values.cell(row=row_idx, column=previous_column_index).value
        
        copy_cell_style(previous_cell_formulas, new_cell_formulas)

        if row_idx == 3:
            new_cell_formulas.value = new_date
        elif row_idx == 4:
            new_cell_formulas.value = new_key_rate / 100 
            new_cell_formulas.number_format = '0.0%' # Excel will add hundredths if necessary
        elif row_idx == 7:
            new_cell_formulas.value = new_rub_usd
        elif row_idx == 8:
            new_cell_formulas.value = new_rub_eur
        elif row_idx == 9:
            new_cell_formulas.value = new_rub_cny
        else:
            if previous_cell_formulas.data_type == 'f':
                translator = Translator(previous_cell_formulas.value, origin=previous_cell_formulas.coordinate)
                new_cell_formulas.value = translator.translate_formula(new_cell_formulas.coordinate)
            else:
                new_cell_formulas.value = previous_cell_formulas.value

        if previous_cell_formulas.value is not None:
            previous_cell_formulas.value = previous_cell_static_value
            
    # Set the column width
    new_col_letter = get_column_letter(new_column_index)
    prev_col_letter = get_column_letter(previous_column_index)
    if sheet_formulas.column_dimensions[prev_col_letter].width:
        sheet_formulas.column_dimensions[new_col_letter].width = sheet_formulas.column_dimensions[prev_col_letter].width
    print("Готово!")